from openpyxl import load_workbook
import json
from datetime import datetime
from db import execute_insert_query

def get_order_data_as_json():
    # 注文リスト☆総務用 (R03.02~）.xlsxファイルを読み込む
    workbook = load_workbook(r'\\fserver\asks\仕入\仕入-国内\毎日の注文\総務専用\注文リスト☆総務用 (R03.02~）.xlsx', data_only=True)

    # 最初のシートを取得
    sheet = workbook.worksheets[0]

    orders = []

    # G列が空白のセルに-を格納して保存
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=7).value is None:
            order_no = sheet.cell(row=row, column=1).value
            order_user = sheet.cell(row=row, column=2).value
            
            # 注文者が記載されている行のみ取得
            if order_user is None:
                continue
            
            order_date = sheet.cell(row=row, column=3).value

            if isinstance(order_date, datetime):
                if order_date.date() == datetime.today().date():
                    order_date = order_date.strftime('%Y-%m-%d')
                else:
                    # order_date = order_date.strftime('%Y-%m-%d')
                    # 注文書作成日が当日のもののみ取得
                    continue
                
            com_no = sheet.cell(row=row, column=10).value
            com_name = sheet.cell(row=row, column=11).value
            name = sheet.cell(row=row, column=17).value
            s_name = sheet.cell(row=row, column=18).value
            quantity = sheet.cell(row=row, column=22).value
            order_unit = sheet.cell(row=row, column=23).value
            deli_location = sheet.cell(row=row, column=29).value

            order_data = {
                "order_no": order_no,
                "order_user": order_user,
                "order_date": order_date,
                "com_no": com_no,
                "com_name": com_name,
                "name": name,
                "s_name": s_name,
                "quantity": quantity,
                "order_unit": order_unit,
                "deli_location": deli_location
            }
            orders.append(order_data)

    return json.dumps(orders, ensure_ascii=False)

if __name__ == '__main__':
    getOrders = get_order_data_as_json()
    print(getOrders)
    
    orders = json.loads(getOrders)
    values = []
    for order in orders:
        value = f"('{order['order_no']}', '{order['order_user']}', '{order['order_date']}', '{order['com_no']}', '{order['com_name']}', '{order['name']}', '{order['s_name']}', '{order['quantity']}', '{order['order_unit']}', '{order['deli_location']}')"
        values.append(value)
    
    sql = f"""
    INSERT INTO initial_orders (order_no, order_user, order_date, com_no, com_name, name, s_name, quantity, order_unit, deli_location)
    VALUES {', '.join(values)}
    """
    print(sql)
    execute_insert_query(sql)
    
